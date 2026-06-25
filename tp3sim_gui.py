import random
import math
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import filedialog


# *----------SIMULACION---------------

def respuesta_individuo(r, p1, p2, p3):
    if r < p1:
        return "Se nego a responder"
    elif r < p1 + p2:
        return "Recordaba el mensaje"
    elif r < p1 + p2 + p3:
        return "No podia recordar el mensaje"

def respuesta_secundaria(r, p1, p2, p3):
    if r < p1:
        return "Definitivamente no"
    elif r < p1 + p2:
        return "Dudoso"
    elif r < p1 + p2 + p3:
        return "Definitivamente si"

def run_simulation(n, p_neg, p_rec, p_norec, p_rec_no, p_rec_dud, p_rec_si, p_norec_no, p_norec_dud, p_norec_si):
    rows = []
    dudoso_acm = 0
    p_dudoso = 0
    for i in range(1, n+1):
        r1 = random.random()
        resp1 = respuesta_individuo(r1, p_neg, p_rec, p_norec)
        if resp1 in ("Recordaba el mensaje", "No podia recordar el mensaje"):
            r2 = random.random()
            if resp1 == "Recordaba el mensaje":
                resp2 = respuesta_secundaria(r2, p_rec_no, p_rec_dud, p_rec_si)
            else:
                resp2 = respuesta_secundaria(r2, p_norec_no, p_norec_dud, p_norec_si)

            if resp2 == "Dudoso":
                dudoso_acm += 1
            p_dudoso = dudoso_acm / i
            rows.append((i, round(r1, 4), resp1, round(r2, 4), resp2, dudoso_acm, p_dudoso))
        else:
            p_dudoso = dudoso_acm / i
            rows.append((i, round(r1, 4), resp1, "", "", dudoso_acm, p_dudoso))
    return rows


# *---------------------INTERFAZ GRAFICA---------------------------------------

COLS = ("Nro", "RAND1", "Respuesta 1", "RAND2", "Respuesta 2", "Dudoso ACM", "P() Dudoso")
COL_WIDTHS = (60, 70, 200, 70, 160, 90, 90)
SEL_BG = "#f5c542"
SEL_FG = "#1a1a1a"


class SimApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Simulación Monte Carlo – tp3sim")
        self.resizable(True, True)
        self.configure(bg="#f0f0f0")

        self._all_rows = []
        self._displayed_rows = []
        self._selected_iids = set()

        self._build_ui()

    def _build_ui(self):

        ctrl = tk.LabelFrame(self, text="Parámetros", bg="#f0f0f0", padx=8, pady=6)
        ctrl.pack(fill="x", padx=10, pady=(10, 4))

        # fila 0 - Ingreso de individuos y limites de muestreo de filas
        tk.Label(ctrl, text="Nro de individuos:", bg="#f0f0f0").grid(row=0, column=0, sticky="e", padx=4)
        self.e_n = tk.Entry(ctrl, width=8); self.e_n.insert(0, "10"); self.e_n.grid(row=0, column=1, sticky="w")

        tk.Label(ctrl, text="Mostrar desde:", bg="#f0f0f0").grid(row=0, column=2, sticky="e", padx=(12,4))
        self.e_from = tk.Entry(ctrl, width=8); self.e_from.insert(0, "1"); self.e_from.grid(row=0, column=3, sticky="w")

        tk.Label(ctrl, text="Hasta:", bg="#f0f0f0").grid(row=0, column=4, sticky="e", padx=(12,4))
        self.e_to = tk.Entry(ctrl, width=8); self.e_to.insert(0, "10"); self.e_to.grid(row=0, column=5, sticky="w")

        # fila 1 - Grupo de probababilidades 1
        tk.Label(ctrl, text="─── Respuesta individuo ───", bg="#f0f0f0", fg="#555").grid(
            row=1, column=0, columnspan=6, sticky="w", padx=4, pady=(8,2))

        labels1 = [("P(Se negó):", "e_p_neg", "0.1"), ("P(Recordaba):", "e_p_rec", "0.4"), ("P(No recordaba):", "e_p_norec", "0.5")]
        for col, (lbl, attr, val) in enumerate(labels1):
            tk.Label(ctrl, text=lbl, bg="#f0f0f0").grid(row=2, column=col*2, sticky="e", padx=4)
            e = tk.Entry(ctrl, width=7); e.insert(0, val); e.grid(row=2, column=col*2+1, sticky="w")
            setattr(self, attr, e)

        # fila 2 - Grupo de probabilidades 2
        tk.Label(ctrl, text="─── Si recordaba el mensaje ───", bg="#f0f0f0", fg="#555").grid(
            row=3, column=0, columnspan=6, sticky="w", padx=4, pady=(8,2))

        labels2 = [("P(Def. no):", "e_rec_no", "0.3"), ("P(Dudoso):", "e_rec_dud", "0.3"), ("P(Def. sí):", "e_rec_si", "0.4")]
        for col, (lbl, attr, val) in enumerate(labels2):
            tk.Label(ctrl, text=lbl, bg="#f0f0f0").grid(row=4, column=col*2, sticky="e", padx=4)
            e = tk.Entry(ctrl, width=7); e.insert(0, val); e.grid(row=4, column=col*2+1, sticky="w")
            setattr(self, attr, e)

        # fila 3 - Grupo de probabilidades 3
        tk.Label(ctrl, text="─── Si NO recordaba el mensaje ───", bg="#f0f0f0", fg="#555").grid(
            row=5, column=0, columnspan=6, sticky="w", padx=4, pady=(8,2))

        labels3 = [("P(Def. no):", "e_norec_no", "0.5"), ("P(Dudoso):", "e_norec_dud", "0.4"), ("P(Def. sí):", "e_norec_si", "0.1")]
        for col, (lbl, attr, val) in enumerate(labels3):
            tk.Label(ctrl, text=lbl, bg="#f0f0f0").grid(row=6, column=col*2, sticky="e", padx=4)
            e = tk.Entry(ctrl, width=7); e.insert(0, val); e.grid(row=6, column=col*2+1, sticky="w")
            setattr(self, attr, e)

        # botones
        btn_frame = tk.Frame(ctrl, bg="#f0f0f0")
        btn_frame.grid(row=7, column=0, columnspan=6, pady=(10,2), sticky="w", padx=4)
        tk.Button(btn_frame, text="▶  Simular", command=self._simulate, bg="#3a7bd5", fg="white",
                  font=("Helvetica", 10, "bold"), relief="flat", padx=12, pady=4).pack(side="left", padx=(0,8))
        tk.Button(btn_frame, text="📥  Exportar a Excel", command=self._export_excel, bg="#217346", fg="white",
                  font=("Helvetica", 10), relief="flat", padx=12, pady=4).pack(side="left")

        # label de resultado
        self.lbl_result = tk.Label(self, text="", bg="#f0f0f0", font=("Helvetica", 10, "italic"), fg="#333")
        self.lbl_result.pack(anchor="w", padx=14, pady=(0,2))

        # tabla de simulacion
        table_frame = tk.Frame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0,10))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=24, font=("Helvetica", 9))
        style.configure("Treeview.Heading", font=("Helvetica", 9, "bold"), background="#3a7bd5", foreground="white")
        style.map("Treeview", background=[("selected", SEL_BG)], foreground=[("selected", SEL_FG)])

        # encabezado
        self.tree = ttk.Treeview(table_frame, columns=COLS, show="headings", selectmode="extended")
        for col, w in zip(COLS, COL_WIDTHS):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center", stretch=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        # Dejar linea pintada
        self.tree.bind("<ButtonRelease-1>", self._on_click)

        # Cambio de color de fondo de lineas
        self.tree.tag_configure("odd",  background="#ffffff")
        self.tree.tag_configure("even", background="#eef2f8")
        self.tree.tag_configure("sel_persist", background=SEL_BG, foreground=SEL_FG)
        self.tree.tag_configure("last", background="#d4edda", foreground="#155724", font=("Helvetica", 9, "bold"))

    # Funciones de verificacion

    def _get_float(self, entry, name):
        try:
            val = float(entry.get().replace(",", "."))
        except ValueError:
            raise ValueError(f"'{name}' debe ser un número.")
        return val

    def _get_int(self, entry, name):
        try:
            val = int(entry.get())
        except ValueError:
            raise ValueError(f"'{name}' debe ser un entero.")
        return val

    def _validate_prob_group(self, vals, group_name):
        for v in vals:
            if v < 0 or v > 1:
                raise ValueError(f"Las probabilidades de '{group_name}' deben ser >= 0 y <= 1.")
            
        total = round(sum(vals), 10)
        if not math.isclose(total, 1.0, abs_tol=1e-9):
            raise ValueError(f"Las probabilidades de '{group_name}' suman {total:.9f}, deben sumar 1.0")

    def _read_params(self):
        n     = self._get_int(self.e_n, "Nro de individuos")
        frm   = self._get_int(self.e_from, "Desde")
        to    = self._get_int(self.e_to, "Hasta")

        p_neg   = self._get_float(self.e_p_neg,   "P(Se negó)")
        p_rec   = self._get_float(self.e_p_rec,   "P(Recordaba)")
        p_norec = self._get_float(self.e_p_norec, "P(No recordaba)")

        p_rec_no  = self._get_float(self.e_rec_no,  "P(Def. no | recordaba)")
        p_rec_dud = self._get_float(self.e_rec_dud, "P(Dudoso | recordaba)")
        p_rec_si  = self._get_float(self.e_rec_si,  "P(Def. sí | recordaba)")

        p_nr_no  = self._get_float(self.e_norec_no,  "P(Def. no | no recordaba)")
        p_nr_dud = self._get_float(self.e_norec_dud, "P(Dudoso | no recordaba)")
        p_nr_si  = self._get_float(self.e_norec_si,  "P(Def. sí | no recordaba)")

        self._validate_prob_group([p_neg, p_rec, p_norec], "Respuesta individuo")
        self._validate_prob_group([p_rec_no, p_rec_dud, p_rec_si], "Si recordaba el mensaje")
        self._validate_prob_group([p_nr_no, p_nr_dud, p_nr_si], "Si NO recordaba el mensaje")

        if n < 1:
            raise ValueError("El número de individuos debe ser ≥ 1.")
        if frm < 1 or to > n or frm > to:
            raise ValueError(f"Rango de visualización inválido (debe ser 1 ≤ desde ≤ hasta ≤ {n}).")

        return n, frm, to, p_neg, p_rec, p_norec, p_rec_no, p_rec_dud, p_rec_si, p_nr_no, p_nr_dud, p_nr_si

    # ── actions ──────────────────────────────────────────────────────────────

    def _simulate(self):
        try:
            n, frm, to, *probs = self._read_params()
        except ValueError as e:
            messagebox.showerror("Error de validación", str(e))
            return

        self._selected_iids.clear()
        self._all_rows = run_simulation(n, *probs)
        last_row = self._all_rows[-1]

        # rows to display: range + always the last row
        idx_range = set(range(frm - 1, to))     # 0-based indices
        idx_range.add(n - 1)                      # always include last
        self._displayed_rows = [self._all_rows[i] for i in sorted(idx_range)]

        self._refresh_table()

        dudoso_prob = last_row[5] / n
        self.lbl_result.config(
            text=f"Probabilidad de respuesta 'Dudoso': {dudoso_prob:.4f}   "
                 f"(Dudoso ACM total: {last_row[5]}  /  {n} individuos simulados)"
        )

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        last_nro = self._all_rows[-1][0] if self._all_rows else None

        for idx, row in enumerate(self._displayed_rows):
            vals = tuple(f"{v:.4f}" if isinstance(v, float) else v for v in row)

            if row[0] == last_nro:
                tag = "last"
            elif row[0] in self._selected_iids:
                tag = "sel_persist"
            else:
                tag = "odd" if idx % 2 == 0 else "even"

            self.tree.insert("", "end", iid=str(row[0]), values=vals, tags=(tag,))

    def _on_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        nro = int(item)
        if nro in self._selected_iids:
            self._selected_iids.discard(nro)
        else:
            self._selected_iids.add(nro)
        self._refresh_table()
        # keep treeview selection in sync
        self.tree.selection_set([str(i) for i in self._selected_iids if self.tree.exists(str(i))])

    def _export_excel(self):
        if not self._displayed_rows:
            messagebox.showwarning("Sin datos", "Primero ejecute la simulación.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar como Excel"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Simulación"

        # header style
        hdr_fill = PatternFill("solid", fgColor="3A7BD5")
        hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align

        # data rows
        last_nro = self._all_rows[-1][0] if self._all_rows else None
        last_fill = PatternFill("solid", fgColor="D4EDDA")
        sel_fill  = PatternFill("solid", fgColor="F5C542")
        even_fill = PatternFill("solid", fgColor="EEF2F8")
        data_font = Font(name="Arial", size=9)

        for row_idx, row in enumerate(self._all_rows, 2):
            vals = [round(v, 4) if isinstance(v, float) else v for v in row]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center")
                if row[0] == last_nro:
                    cell.fill = last_fill
                    cell.font = Font(name="Arial", size=9, bold=True)
                elif row[0] in self._selected_iids:
                    cell.fill = sel_fill
                elif (row_idx - 2) % 2 == 1:
                    cell.fill = even_fill

        # column widths
        for col, width in zip(ws.columns, [10, 10, 28, 10, 22, 14, 12, 12]):
            ws.column_dimensions[col[0].column_letter].width = width

        # freeze header row
        ws.freeze_panes = "A2"

        wb.save(path)
        messagebox.showinfo("Exportado", f"Archivo guardado en:\n{path}")


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = SimApp()
    app.mainloop()
